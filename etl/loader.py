from __future__ import annotations

import json
import re
import unicodedata
from copy import deepcopy
from datetime import date, datetime
import os
import sys
from pathlib import Path
from typing import Any, Iterable
from functools import wraps

import pandas as pd
import openpyxl

from config.settings import (
    PROFIT_ADVANCE_ADJUSTMENT_PCT,
    PROFIT_ADVANCE_ADJUSTMENT_PER_QUOTA,
    PROFIT_ADVANCE_MONTHLY_RESULT,
    PROFIT_ADVANCE_PAYMENT,
    PROFIT_ADVANCE_REFERENCE,
    PROFIT_ADVANCE_TOTAL_ADJUSTMENT,
    PROFIT_ADVANCE_TOTAL_QUOTAS,
)
from utils.analysis_generator import generate_insights, generate_technical_analysis
from utils.dashboard_metrics import build_dashboard_metrics
from utils.number_formatter import formatar_valor_monetario

def _runtime_root() -> Path:
    env_root = os.getenv("DASHBOARD_APP_ROOT")
    if env_root:
        return Path(env_root)

    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent.parent


DATA_DIR = _runtime_root() / "data"
TEMPLATE_FILE = DATA_DIR / "dashboard_content.json"
ORCAMENTO_FILE = DATA_DIR / "Orçamento.xlsx"
FOLHA_FILE = DATA_DIR / "Resumo - folha.xlsx"
ANTECIPACAO_FILE = DATA_DIR / "INFORMAÇÕES GERENCIAIS.xlsx"
PERFORMANCE_FILE = DATA_DIR / "PERFORMANCE.XLSX"
OMIE_COSTS_FILE = DATA_DIR / "Custos OMIE.xlsx"

MONTH_LABELS = {
    1: "jan", 2: "fev", 3: "mar", 4: "abr", 
    5: "mai", 6: "jun", 7: "jul", 8: "ago", 
    9: "set", 10: "out", 11: "nov", 12: "dez"
}

MONTH_SHEET_NAMES = {
    1: "janeiro",
    2: "fevereiro",
    3: "marco",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro",
}

OMIE_COST_SHEET_NAMES = {
    1: "Custos jan26",
    2: "Custos fev26",
    3: "Custos mar26",
    4: "Custos abr26",
    5: "Custos mai26",
    6: "Custos jun26",
    7: "Custos jul26",
    8: "Custos ago26",
    9: "Custos set26",
    10: "Custos out26",
    11: "Custos nov26",
    12: "Custos dez26",
}

CLIENT_COST_SHEETS = [
    "Bradesco", "Santander", "MadeiraMadeira", "Claro", "UOL", "Vivo",
    "BS2", "Naturgy", "Mercado Livre", "Daycoval", "Enel", "MRS", "XS4",
    "Banco PAN", "NIO", "Quinto Andar", "Patrimônio", "Resolução de Disputas",
    "Vero", "Criminal", "Recupera", "Control Jurídica", "Pauta", "Saneamento",
]


def cache_data(*_args, **_kwargs):
    """No-op cache decorator kept for compatibility with the ETL API."""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            return func(*args, **kwargs)

        return wrapper

    return decorator


def _to_float(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = text.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    text = re.sub(r"[^0-9\-\.]+", "", text)
    if text in {"", "-", "."}:
        return 0.0

    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_valid_excel_date(value: Any) -> bool:
    return isinstance(value, datetime)


def _is_january_to_selected_months(selected_months: list[int]) -> bool:
    months = sorted({month for month in selected_months if 1 <= month <= 12})
    return bool(months) and months == list(range(1, max(months) + 1))


def load_performance_budget_metrics(file_path: str, year: int, selected_months: list[int]) -> dict[str, Any]:
    """
    Le a aba Performance usando apenas a linha consolidada 1 - TODOS e retorna
    os valores orcados do periodo filtrado.
    """
    months = sorted({month for month in selected_months if 1 <= month <= 12})
    result = {
        "budgetRevenue": 0.0,
        "budgetCosts": 0.0,
        "budgetResult": 0.0,
        "source": "PERFORMANCE.XLSX",
        "budgetLogic": "Linha consolidada 1 - TODOS; soma mensal K/O/AB conforme meses selecionados",
        "months": months,
        "validation": {
            "isJanuaryToPeriod": _is_january_to_selected_months(months),
            "revenueAccumulated": 0.0,
            "costsAccumulated": 0.0,
            "resultAccumulated": 0.0,
            "revenueDiff": 0.0,
            "costsDiff": 0.0,
            "resultDiff": 0.0,
        },
        "error": None,
    }

    path = Path(file_path)
    if not path.exists():
        result["error"] = f"Arquivo nao encontrado: {path.name}"
        print(f"[ETL Performance] {result['error']}")
        return result

    if not months:
        result["error"] = "Nenhum mes valido selecionado."
        print(f"[ETL Performance] {result['error']}")
        return result

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = workbook["Performance"]
    except Exception as exc:
        result["error"] = f"Nao foi possivel ler Performance: {exc}"
        print(f"[ETL Performance] {result['error']}")
        return result

    last_accumulated: dict[str, float] = {}
    last_month = max(months)
    matching_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        group = _normalize_text(row[1] if len(row) > 1 else "")
        carteira = _normalize_text(row[3] if len(row) > 3 else "")
        if group != "1 - todos" and carteira != "1 - todos":
            continue

        month_date = row[7] if len(row) > 7 else None
        if not _is_valid_excel_date(month_date):
            continue
        if month_date.year != year or month_date.month not in months:
            continue

        matching_rows += 1
        result["budgetRevenue"] += _to_float(row[10] if len(row) > 10 else 0.0)
        result["budgetCosts"] += _to_float(row[14] if len(row) > 14 else 0.0)
        result["budgetResult"] += _to_float(row[27] if len(row) > 27 else 0.0)

        if month_date.month == last_month:
            last_accumulated = {
                "revenueAccumulated": _to_float(row[11] if len(row) > 11 else 0.0),
                "costsAccumulated": _to_float(row[15] if len(row) > 15 else 0.0),
                "resultAccumulated": _to_float(row[28] if len(row) > 28 else 0.0),
            }

    if matching_rows == 0:
        result["error"] = "Linha consolidada 1 - TODOS nao encontrada para o periodo selecionado."
        print(f"[ETL Performance] {result['error']}")
        return result

    if result["validation"]["isJanuaryToPeriod"] and last_accumulated:
        result["validation"].update(last_accumulated)
        result["validation"]["revenueDiff"] = result["budgetRevenue"] - last_accumulated["revenueAccumulated"]
        result["validation"]["costsDiff"] = result["budgetCosts"] - last_accumulated["costsAccumulated"]
        result["validation"]["resultDiff"] = result["budgetResult"] - last_accumulated["resultAccumulated"]

    return result


def load_people_costs_from_omie(file_path: str, selected_months: list[int]) -> dict[str, Any]:
    """Le Custos OMIE.xlsx e retorna os custos financeiros de pessoas por periodo."""
    months = sorted({month for month in selected_months if 1 <= month <= 12})
    result = {
        "partners": 0.0,
        "clt": 0.0,
        "totalPeopleCost": 0.0,
        "breakdown": [
            {"name": "Sócios de Serviço e Capital", "value": 0.0},
            {"name": "Celetistas", "value": 0.0},
        ],
        "source": "Custos OMIE.xlsx",
        "months": [],
        "available": False,
        "error": None,
    }

    path = Path(file_path)
    if not path.exists():
        result["error"] = f"Arquivo nao encontrado: {path.name}"
        print(f"[ETL OMIE] {result['error']}")
        return result

    if not months:
        result["error"] = "Nenhum mes valido selecionado."
        print(f"[ETL OMIE] {result['error']}")
        return result

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result["error"] = f"Nao foi possivel ler Custos OMIE: {exc}"
        print(f"[ETL OMIE] {result['error']}")
        return result

    for month in months:
        sheet_name = OMIE_COST_SHEET_NAMES.get(month)
        if not sheet_name or sheet_name not in workbook.sheetnames:
            print(f"[ETL OMIE] Aba nao encontrada para mes {month}: {sheet_name}")
            continue

        ws = workbook[sheet_name]
        result["months"].append(month)
        for row in ws.iter_rows(min_row=2, values_only=True):
            label = _normalize_text(row[0] if len(row) > 0 else "")
            value = _to_float(row[1] if len(row) > 1 else 0.0)
            if label == _normalize_text("Sócios de Serviço e Capital"):
                result["partners"] += value
            elif label == _normalize_text("Celetistas"):
                result["clt"] += value

    result["totalPeopleCost"] = result["partners"] + result["clt"]
    result["breakdown"] = [
        {"name": "Sócios de Serviço e Capital", "value": result["partners"]},
        {"name": "Celetistas", "value": result["clt"]},
    ]
    result["available"] = bool(result["months"]) and result["totalPeopleCost"] > 0
    if not result["available"] and result["error"] is None:
        result["error"] = "Nenhum custo financeiro de pessoas encontrado na OMIE para o periodo."
        print(f"[ETL OMIE] {result['error']}")

    return result


def _format_currency(value: float) -> str:
    return formatar_valor_monetario(value)


def _format_currency_full(value: float) -> str:
    return _format_currency(value)


def _format_currency_precise(value: float) -> str:
    sign = "-" if value < 0 else ""
    formatted = f"{abs(float(value)):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {sign}{formatted}"


def _format_percent(value: float) -> str:
    return f"{value:.1f}%".replace(".", ",")


def _format_signed_percent(value: float) -> str:
    sign = "+" if value >= 0 else ""
    return f"{sign}{_format_percent(value * 100)}"


def _format_count(value: float) -> str:
    return str(int(round(float(value))))


def _format_plain_decimal(value: float, decimals: int = 2) -> str:
    return f"{float(value):,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_months_label(months: list[int], year: int) -> str:
    if not months:
        return f"{year}"

    month_labels = [MONTH_LABELS.get(month, str(month)) for month in months]
    year_suffix = str(year)[-2:]
    if len(month_labels) == 1:
        return f"{month_labels[0]}/{year_suffix}"
    return f"{', '.join(month_labels[:-1])} e {month_labels[-1]}/{year_suffix}"


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().lower()


def _month_number_from_sheet_name(sheet_name: str) -> int | None:
    normalized = _normalize_text(sheet_name)
    for month, name in MONTH_SHEET_NAMES.items():
        if normalized == name:
            return month
    return None


@cache_data(show_spinner=False)
def _load_dashboard_template(file_mtime_ns: int) -> dict[str, Any]:
    return json.loads(TEMPLATE_FILE.read_text(encoding="utf-8"))


def _is_date_in_period(col_val, ano, selected_months):
    """Verifica se a data do Excel pertence ao ano e meses selecionados."""
    if pd.isna(col_val):
        return False
    try:
        if isinstance(col_val, (int, float)):
            # Formato numérico do Excel
            d = pd.to_datetime(col_val, origin="1899-12-30", unit="D")
        else:
            d = pd.to_datetime(col_val)
        return d.year == ano and d.month in selected_months
    except Exception:
        return False


def _sum_row_for_period(df: pd.DataFrame, label: str, header_row: int, ano: int, selected_months: list[int]) -> float:
    expected = _normalize_text(label)
    for i in range(len(df)):
        if _normalize_text(df.iloc[i, 0]) == expected:
            return sum(
                _to_float(df.iloc[i, j])
                for j in range(1, len(df.columns))
                if _is_date_in_period(df.iloc[header_row, j], ano, selected_months)
            )
    return 0.0


def _load_totais_period(selected_months: list[int], year: int = 2026) -> dict[str, float]:
    """Carrega os totais oficiais do BI na aba TOTAIS  20 E 21."""
    empty = {
        "receita": 0.0,
        "custos": 0.0,
        "resultado": 0.0,
        "impostos": 0.0,
        "sucumbencia": 0.0,
        "receita_2025": 0.0,
        "custos_2025": 0.0,
        "resultado_2025": 0.0,
        "variacao_receita": 0.0,
        "variacao_custos": 0.0,
        "official_months": [],
    }
    if not ANTECIPACAO_FILE.exists():
        return empty

    try:
        wb = openpyxl.load_workbook(
            ANTECIPACAO_FILE,
            read_only=True,
            data_only=True,
        )
        ws = wb["TOTAIS  20 E 21"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        print(f"[ETL] Nao foi possivel ler TOTAIS  20 E 21: {exc}")
        return empty

    header = rows[0]
    period_cols = []
    previous_cols = []
    for index, header_value in enumerate(header):
        if not hasattr(header_value, "year"):
            continue
        if header_value.year == year and header_value.month in selected_months:
            period_cols.append(index)
        if header_value.year == year - 1 and header_value.month in selected_months:
            previous_cols.append(index)

    print(f"Periodo selecionado {year} meses {selected_months}:")
    print(f"  Colunas {year}: {[header[index].strftime('%b/%Y') for index in period_cols]}")
    print(f"  Colunas {year - 1}: {[header[index].strftime('%b/%Y') for index in previous_cols]}")

    def sum_row(label: str, columns: list[int]) -> float:
        expected = _normalize_text(label)
        for row in rows[1:]:
            if _normalize_text(row[0]) == expected:
                values = [row[index] for index in columns if row[index] not in (None, "", "-")]
                return sum(float(value) for value in values if isinstance(value, (int, float)))
        return 0.0

    def row_value(label: str, column: int) -> float:
        expected = _normalize_text(label)
        for row in rows[1:]:
            if _normalize_text(row[0]) == expected:
                value = row[column]
                return float(value) if isinstance(value, (int, float)) else 0.0
        return 0.0

    receita = sum_row("receita", period_cols)
    custos = sum_row("custos e despesas", period_cols)
    resultado = sum_row("resultado ig", period_cols)
    impostos = sum_row("impostos", period_cols)
    sucumbencia = sum_row("sucumbencia", period_cols)
    receita_2025 = sum_row("receita", previous_cols)
    custos_2025 = sum_row("custos e despesas", previous_cols)
    resultado_2025 = sum_row("resultado ig", previous_cols)
    official_months = [
        {
            "month": header[index].month,
            "receita": row_value("receita", index),
            "custos": row_value("custos e despesas", index),
            "resultado": row_value("resultado ig", index),
        }
        for index in period_cols
    ]

    return {
        "receita": receita,
        "custos": custos,
        "resultado": resultado,
        "impostos": impostos,
        "sucumbencia": sucumbencia,
        "receita_2025": receita_2025,
        "custos_2025": custos_2025,
        "resultado_2025": resultado_2025,
        "variacao_receita": ((receita - receita_2025) / receita_2025) if receita_2025 else 0.0,
        "variacao_custos": ((custos - custos_2025) / custos_2025) if custos_2025 else 0.0,
        "official_months": official_months,
    }


def _month_columns_from_header(df: pd.DataFrame, header_row: int, ano: int, selected_months: list[int]) -> list[int]:
    columns = [
        j
        for j in range(1, min(13, len(df.columns)))
        if _is_date_in_period(df.iloc[header_row, j], ano, selected_months)
    ]
    if columns:
        return columns

    return [
        month
        for month in selected_months
        if 1 <= month <= 12 and month < len(df.columns)
    ]


def _sum_fixed_row(df: pd.DataFrame, row_index: int, columns: Iterable[int]) -> float:
    if row_index >= len(df):
        return 0.0
    return sum(_to_float(df.iloc[row_index, column]) for column in columns)


def _row_text(df: pd.DataFrame, row_index: int, columns: Iterable[int] = (0, 1)) -> str:
    parts = []
    for column in columns:
        if column < len(df.columns):
            text = str(df.iloc[row_index, column]).strip()
            if text and text.lower() != "nan":
                parts.append(text)
    return " ".join(parts)


def _is_month_label(value: Any, selected_months: list[int]) -> bool:
    normalized = _normalize_text(value)
    selected_labels = {MONTH_LABELS[month] for month in selected_months if month in MONTH_LABELS}
    return normalized in selected_labels


def _get_fat_real_cols(df: pd.DataFrame, header_row: int, selected_months: list[int]) -> list[int]:
    month_row = header_row - 1
    if month_row < 0:
        return []

    month_starts = [
        column
        for column in range(len(df.columns))
        if _is_month_label(df.iloc[month_row, column], selected_months)
    ]
    if not month_starts:
        return []

    all_month_starts = [
        column
        for column in range(len(df.columns))
        if _normalize_text(df.iloc[month_row, column]) in set(MONTH_LABELS.values())
    ]
    columns = []
    for month_start in month_starts:
        next_months = [column for column in all_month_starts if column > month_start]
        block_end = min(next_months) if next_months else len(df.columns)
        for column in range(month_start, block_end):
            if _normalize_text(df.iloc[header_row, column]) in {"fat real", "faturamento real"}:
                columns.append(column)
                break

    return columns


def _get_valor_real_cost_cols(df: pd.DataFrame, header_row: int, selected_months: list[int]) -> list[int]:
    month_row = header_row - 1
    if month_row < 0:
        return []

    month_starts = [
        column
        for column in range(len(df.columns))
        if _is_month_label(df.iloc[month_row, column], selected_months)
    ]
    if not month_starts:
        return []

    all_month_starts = [
        column
        for column in range(len(df.columns))
        if _normalize_text(df.iloc[month_row, column]) in set(MONTH_LABELS.values())
    ]
    columns = []
    for month_start in month_starts:
        next_months = [column for column in all_month_starts if column > month_start]
        block_end = min(next_months) if next_months else len(df.columns)
        found_column = None
        for label_row in range(header_row + 1, min(header_row + 4, len(df))):
            for column in range(month_start, block_end):
                if _normalize_text(df.iloc[label_row, column]) == "valor real":
                    found_column = column
                    break
            if found_column is not None:
                break
        if found_column is not None:
            columns.append(found_column)

    return columns


def _get_cost_header_rows(df: pd.DataFrame) -> list[int]:
    return [
        i
        for i in range(len(df))
        if any(
            _normalize_text(df.iloc[i, column]) in {"fat real", "faturamento real"}
            for column in range(len(df.columns))
        )
    ]


def _get_cost_block_end(df: pd.DataFrame, start_row: int, next_header_row: int | None) -> int:
    end = next_header_row if next_header_row is not None else len(df)
    for row_index in range(start_row + 1, end):
        label = _normalize_text(_row_text(df, row_index))
        if "total despesas diretas" in label or "total receita" in label:
            return row_index
    return end


def _is_cost_row_excluded(label: str) -> bool:
    if not label or label == "nan":
        return True
    excluir_pessoal = [
        "remuneracao",
        "socio",
        "socios",
        "celetista",
        "estagiario",
        "estagiarios",
        "plano de saude",
        "vale alimentacao",
        "vale transporte",
        "seguro de vida",
        "seg vida",
        "encargos",
        "ferias",
        "13o",
        "plano odontologico",
        "exames medicos",
        "ir s/ salario",
        "indenizacao",
        "bonificacoes",
        "premiacoes",
    ]
    excluded_fragments = [
        # totais e receita
        "total",
        "receita",
        "margem",
        "resultado",
        "rateio",
        "faturamento",
        "distribuicao",
        "distribuição",
        "ajuste mensal",
        "calculo aproximado",
        "cálculo aproximado",
        "margem de contribuicao",
        "margem de contribuição",
        # pessoal - ja contado em Socios/CLT
        "remuneração",
        "remuneracao",
        "13º",
        "13o",
        "ferias",
        "férias",
        "encargos",
        "plano de saude",
        "plano de saúde",
        "plano odontologico",
        "plano odontológico",
        "vale alimentacao",
        "vale alimentação",
        "vale refeicao",
        "vale refeição",
        "vale transporte",
        "seg vida",
        "seguro de vida",
        "exames medicos",
        "exames médicos",
        "ir s/ salario",
        "ir s/ salário",
        "ir - salario",
        "ir - salário",
        "salario",
        "salário",
        "salarios",
        "salários",
        "indenizacao",
        "indenização",
        "encargos rescisorios",
        "encargos rescisórios",
        "bonificacoes",
        "bonificações",
        "premiações",
        "premiacoes",
        "contratacao",
        "contratação",
        # distribuicao de lucros
        "distribuicao anual",
        "distribuição anual",
        "ajuste mensal de lucros",
        # ja possuem campo separado no card
        "impostos",
        "correspondente",
        "correspondentes",
    ]
    excluded_fragments.extend(excluir_pessoal)
    return any(fragment in label for fragment in excluded_fragments)


def _extract_client_costs(
    workbook_path: str,
    selected_months: list[int],
    sheets: Iterable[str] = CLIENT_COST_SHEETS,
) -> dict[str, float]:
    workbook = pd.ExcelFile(workbook_path)
    should_log_other_details = sorted(selected_months) == [1, 2, 3]
    totals = {
        "correspondentes": 0.0,
        "outras_despesas": 0.0,
        "impostos": 0.0,
    }
    other_expense_entries: list[tuple[str, str, float]] = []

    for sheet in sheets:
        if sheet not in workbook.sheet_names:
            print(f"[ETL Custos] {sheet}: aba nao encontrada")
            continue

        df = pd.read_excel(workbook_path, sheet_name=sheet, header=None, engine="openpyxl").fillna("")
        header_rows = _get_cost_header_rows(df)
        sheet_totals = {
            "correspondentes": 0.0,
            "outras_despesas": 0.0,
            "impostos": 0.0,
        }
        sheet_other_entries: list[tuple[str, float]] = []

        for header_index, header_row in enumerate(header_rows):
            valor_real_cols = _get_valor_real_cost_cols(df, header_row, selected_months)
            if not valor_real_cols:
                continue

            next_header_row = header_rows[header_index + 1] if header_index + 1 < len(header_rows) else None
            block_end = _get_cost_block_end(df, header_row, next_header_row)

            for row_index in range(header_row + 4, block_end):
                raw_label = _row_text(df, row_index, columns=(0,)) or _row_text(df, row_index, columns=(1,))
                label = _normalize_text(raw_label)
                value = _sum_fixed_row(df, row_index, valor_real_cols)
                if not value:
                    continue

                if "correspondente" in label:
                    sheet_totals["correspondentes"] += value
                elif label == "impostos" or label.startswith("impostos "):
                    sheet_totals["impostos"] += value
                elif not _is_cost_row_excluded(label):
                    sheet_totals["outras_despesas"] += value
                    if should_log_other_details:
                        clean_label = re.sub(r"\s+", " ", str(raw_label)).strip()
                        sheet_other_entries.append((clean_label, value))
                        other_expense_entries.append((sheet, clean_label, value))

        for key in totals:
            totals[key] += sheet_totals[key]

        print(
            "[ETL Custos] "
            f"{sheet}: correspondentes={sheet_totals['correspondentes']:.2f}; "
            f"outras={sheet_totals['outras_despesas']:.2f}; "
            f"impostos={sheet_totals['impostos']:.2f}"
        )
        if should_log_other_details and sheet_other_entries:
            print(f"[ETL Custos] {sheet}: linhas somadas em outras_despesas")
            for label, value in sheet_other_entries:
                print(f"    - {label}: {value:.2f}")

    totals["socios_servico"] = 0.0
    totals["clt"] = 0.0
    totals["total"] = (
        totals["correspondentes"]
        + totals["outras_despesas"]
        + totals["impostos"]
        + totals["socios_servico"]
        + totals["clt"]
    )
    print(
        "[ETL Custos] TOTAL: "
        f"correspondentes={totals['correspondentes']:.2f}; "
        f"outras={totals['outras_despesas']:.2f}; "
        f"impostos={totals['impostos']:.2f}; "
        f"total={totals['total']:.2f}"
    )
    if should_log_other_details:
        if 200_000 <= totals["outras_despesas"] <= 500_000:
            print(
                "[ETL Custos] VALIDACAO outras_despesas jan-mar/26 dentro do intervalo: "
                f"{totals['outras_despesas']:.2f}"
            )
        else:
            print(
                "[ETL Custos] ATENCAO outras_despesas jan-mar/26 fora do intervalo esperado: "
                f"{totals['outras_despesas']:.2f}"
            )
            print("[ETL Custos] Top 10 linhas que mais inflaram outras_despesas:")
            for sheet, label, value in sorted(other_expense_entries, key=lambda item: abs(item[2]), reverse=True)[:10]:
                print(f"    - {sheet} | {label}: {value:.2f}")
    return totals


def _get_row_total_value(df: pd.DataFrame, label: str) -> float:
    expected = _normalize_text(label)
    for i in range(len(df)):
        if _normalize_text(df.iloc[i, 0]) == expected:
            for j in range(13, 0, -1):
                value = _to_float(df.iloc[i, j])
                if value:
                    return value
    return 0.0


def _get_table_value_by_labels(
    df: pd.DataFrame,
    row_label: str,
    column_label: str,
    header_row: int = 2,
) -> float:
    expected_row = _normalize_text(row_label)
    expected_col = _normalize_text(column_label)
    column_index = None

    for j in range(len(df.columns)):
        if _normalize_text(df.iloc[header_row, j]) == expected_col:
            column_index = j
            break

    if column_index is None:
        return 0.0

    for i in range(header_row + 1, len(df)):
        if _normalize_text(df.iloc[i, 0]) == expected_row:
            return _to_float(df.iloc[i, column_index])

    return 0.0


def _get_comparativo_2025_period(df: pd.DataFrame, selected_months: list[int]) -> float:
    """Busca o Realizado 2025 no bloco Receita - Com Sucumbencia.

    A aba Comparativo 26-25 do arquivo atual nao possui colunas mensais; por isso
    usamos o total anual da linha Total - Gondim proporcional ao periodo filtrado.
    Se a planilha ganhar granularidade mensal no futuro, este ponto fica isolado.
    """
    title_row = None
    title_col = None
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if _normalize_text(df.iloc[i, j]) == "receita - com sucumbencia":
                title_row = i
                title_col = j
                break
        if title_row is not None:
            break

    if title_row is None or title_col is None:
        return 0.0

    header_row = title_row + 1
    label_col = None
    value_col = None
    for j in range(title_col, len(df.columns)):
        header = _normalize_text(df.iloc[header_row, j])
        if header == "carteira":
            label_col = j
        elif header == "realizado 2025":
            value_col = j

    if label_col is None or value_col is None:
        return 0.0

    annual_total = 0.0
    for i in range(header_row + 1, len(df)):
        label = _normalize_text(df.iloc[i, label_col])
        if label == "total - gondim":
            annual_total = _to_float(df.iloc[i, value_col])
            break

    if not annual_total:
        values = []
        for i in range(header_row + 1, len(df)):
            label = _normalize_text(df.iloc[i, label_col])
            if not label or label.startswith("total"):
                continue
            values.append(_to_float(df.iloc[i, value_col]))
        annual_total = sum(values)

    month_count = len({month for month in selected_months if 1 <= month <= 12})
    return annual_total * (month_count / 12) if month_count else 0.0


def _get_comparativo_total_2025_period(
    df: pd.DataFrame,
    section_title: str,
    selected_months: list[int],
) -> float:
    """Busca o Realizado 2025 da linha Total - Gondim em um bloco do Comparativo 26-25."""
    expected_title = _normalize_text(section_title)
    title_row = None
    title_col = None
    for i in range(len(df)):
        for j in range(len(df.columns)):
            if _normalize_text(df.iloc[i, j]) == expected_title:
                title_row = i
                title_col = j
                break
        if title_row is not None:
            break

    if title_row is None or title_col is None:
        return 0.0

    header_row = title_row + 1
    label_col = None
    value_col = None
    for j in range(max(title_col - 1, 0), len(df.columns)):
        header = _normalize_text(df.iloc[header_row, j])
        if header == "carteira":
            label_col = j
        elif header == "realizado 2025":
            value_col = j

        if label_col is not None and value_col is not None and j > value_col:
            break

    if label_col is None or value_col is None:
        return 0.0

    annual_total = 0.0
    for i in range(header_row + 1, len(df)):
        label = _normalize_text(df.iloc[i, label_col])
        if label == "total - gondim":
            annual_total = _to_float(df.iloc[i, value_col])
            break

    month_count = len({month for month in selected_months if 1 <= month <= 12})
    return annual_total * (month_count / 12) if month_count else 0.0


def _share(value: float, total: float) -> str:
    return _format_percent((value / total * 100) if total else 0.0)


FOLHA_SOCIOS_CAPITAL = {_normalize_text("Socios de capital GAN")}
FOLHA_SOCIOS_SERVICO = {
    _normalize_text("Socios de servico GAN"),
    _normalize_text("Socios de servico GAT"),
    _normalize_text("Socios de servico GAA"),
}
FOLHA_CLT = {_normalize_text("Celetista")}
FOLHA_ESTAGIARIOS = {
    _normalize_text("Estagiarios RPA + RPA + PJ"),
    _normalize_text("Estagiarios GAN"),
    _normalize_text("Estagiarios GAT"),
}


def _empty_people_and_payroll_data() -> dict[str, Any]:
    return {
        "people": {
            "total": 0,
            "socios": {"qtd": 0, "pct": 0.0},
            "clt": {"qtd": 0, "pct": 0.0},
            "estagiarios": {"qtd": 0, "pct": 0.0},
        },
        "socios_servico_val": 0.0,
        "socios_capital_val": 0.0,
        "socios_val": 0.0,
        "clt_val": 0.0,
        "estagiarios_val": 0.0,
        "total_pessoas_val": 0.0,
        "total_folha_val": 0.0,
        "folha_months_used": [],
        "folha_snapshot_month": None,
    }


def _find_folha_columns(df: pd.DataFrame) -> tuple[int, int, int, int] | None:
    for row_index in range(len(df)):
        headers = {
            _normalize_text(df.iloc[row_index, column_index]): column_index
            for column_index in range(len(df.columns))
        }
        if {"categoria", "quantidade", "valor"}.issubset(headers):
            return row_index, headers["categoria"], headers["quantidade"], headers["valor"]
    return None


def _read_folha_month(sheet_name: str) -> dict[str, Any]:
    df = pd.read_excel(FOLHA_FILE, sheet_name=sheet_name, header=None, engine="openpyxl").fillna("")
    columns = _find_folha_columns(df)
    if columns is None:
        print(f"[ETL] Aba de folha ignorada sem cabecalho esperado: {sheet_name}")
        return {"records": {}, "total_qtd": 0.0, "total_valor": 0.0}

    header_row, category_col, quantity_col, value_col = columns
    records: dict[str, dict[str, float]] = {}
    total_qtd = 0.0
    total_valor = 0.0

    for row_index in range(header_row + 1, len(df)):
        category = _normalize_text(df.iloc[row_index, category_col])
        if not category or category in {"total", "totais"}:
            continue

        qtd = _to_float(df.iloc[row_index, quantity_col])
        valor = _to_float(df.iloc[row_index, value_col])
        records[category] = {"qtd": qtd, "valor": valor}
        total_qtd += qtd
        total_valor += valor

    return {"records": records, "total_qtd": total_qtd, "total_valor": total_valor}


def _sum_folha_categories(month_data: dict[str, Any], categories: set[str], field: str) -> float:
    records = month_data.get("records", {})
    return sum(_to_float(records.get(category, {}).get(field, 0.0)) for category in categories)


def _load_people_and_payroll_data(selected_months: list[int]) -> dict[str, Any]:
    """Le pessoas e valores de folha em data/Resumo - folha.xlsx."""
    if not FOLHA_FILE.exists():
        return _empty_people_and_payroll_data()

    valid_selected_months = sorted({month for month in selected_months if 1 <= month <= 12})
    if not valid_selected_months:
        return _empty_people_and_payroll_data()

    workbook = pd.ExcelFile(FOLHA_FILE, engine="openpyxl")
    monthly_sheets = {
        month: sheet_name
        for sheet_name in workbook.sheet_names
        if (month := _month_number_from_sheet_name(sheet_name)) is not None
    }
    months_to_read = [month for month in valid_selected_months if month in monthly_sheets]

    if not months_to_read:
        fallback_sheet = "Planilha1" if "Planilha1" in workbook.sheet_names else workbook.sheet_names[0]
        print(
            "[ETL] Nenhuma aba mensal de folha encontrada para meses "
            f"{valid_selected_months}; usando {fallback_sheet} como snapshot de quantidade."
        )
        snapshot_data = _read_folha_month(fallback_sheet)
        socios_servico_qtd = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_SERVICO, "qtd")
        socios_capital_qtd = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_CAPITAL, "qtd")
        clt_qtd = _sum_folha_categories(snapshot_data, FOLHA_CLT, "qtd")
        estagiarios_qtd = _sum_folha_categories(snapshot_data, FOLHA_ESTAGIARIOS, "qtd")
        total_profissionais = snapshot_data["total_qtd"]

        def pct_snapshot(value: float) -> float:
            return (value / total_profissionais) if total_profissionais else 0.0

        socios_servico_val = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_SERVICO, "valor")
        socios_capital_val = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_CAPITAL, "valor")
        clt_val = _sum_folha_categories(snapshot_data, FOLHA_CLT, "valor")
        estagiarios_val = _sum_folha_categories(snapshot_data, FOLHA_ESTAGIARIOS, "valor")
        total_pessoas_val = socios_servico_val + socios_capital_val + clt_val + estagiarios_val

        return {
            "people": {
                "total": int(round(total_profissionais)),
                "socios": {"qtd": int(round(socios_servico_qtd)), "pct": pct_snapshot(socios_servico_qtd)},
                "clt": {"qtd": int(round(clt_qtd)), "pct": pct_snapshot(clt_qtd)},
                "estagiarios": {"qtd": int(round(estagiarios_qtd)), "pct": pct_snapshot(estagiarios_qtd)},
            },
            "socios_servico_val": socios_servico_val,
            "socios_capital_val": socios_capital_val,
            "socios_val": socios_servico_val,
            "clt_val": clt_val,
            "estagiarios_val": estagiarios_val,
            "total_pessoas_val": total_pessoas_val,
            "total_folha_val": _to_float(snapshot_data["total_valor"]),
            "folha_months_used": [],
            "folha_snapshot_month": None,
            "socios_capital_qtd": int(round(socios_capital_qtd)),
        }

    monthly_data = {
        month: _read_folha_month(monthly_sheets[month])
        for month in months_to_read
    }
    snapshot_month = max(monthly_data)
    snapshot_data = monthly_data[snapshot_month]

    socios_servico_qtd = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_SERVICO, "qtd")
    socios_capital_qtd = _sum_folha_categories(snapshot_data, FOLHA_SOCIOS_CAPITAL, "qtd")
    clt_qtd = _sum_folha_categories(snapshot_data, FOLHA_CLT, "qtd")
    estagiarios_qtd = _sum_folha_categories(snapshot_data, FOLHA_ESTAGIARIOS, "qtd")
    total_profissionais = snapshot_data["total_qtd"]

    socios_servico_val = sum(
        _sum_folha_categories(month_data, FOLHA_SOCIOS_SERVICO, "valor")
        for month_data in monthly_data.values()
    )
    socios_capital_val = sum(
        _sum_folha_categories(month_data, FOLHA_SOCIOS_CAPITAL, "valor")
        for month_data in monthly_data.values()
    )
    clt_val = sum(
        _sum_folha_categories(month_data, FOLHA_CLT, "valor")
        for month_data in monthly_data.values()
    )
    estagiarios_val = sum(
        _sum_folha_categories(month_data, FOLHA_ESTAGIARIOS, "valor")
        for month_data in monthly_data.values()
    )
    total_folha_val = sum(_to_float(month_data["total_valor"]) for month_data in monthly_data.values())
    total_pessoas_val = socios_servico_val + socios_capital_val + clt_val + estagiarios_val

    def pct(value: float) -> float:
        return (value / total_profissionais) if total_profissionais else 0.0

    return {
        "people": {
            "total": int(round(total_profissionais)),
            "socios": {"qtd": int(round(socios_servico_qtd)), "pct": pct(socios_servico_qtd)},
            "clt": {"qtd": int(round(clt_qtd)), "pct": pct(clt_qtd)},
            "estagiarios": {"qtd": int(round(estagiarios_qtd)), "pct": pct(estagiarios_qtd)},
        },
        "socios_servico_val": socios_servico_val,
        "socios_capital_val": socios_capital_val,
        "socios_val": socios_servico_val,
        "clt_val": clt_val,
        "estagiarios_val": estagiarios_val,
        "total_pessoas_val": total_pessoas_val,
        "total_folha_val": total_folha_val,
        "folha_months_used": months_to_read,
        "folha_snapshot_month": snapshot_month,
        "socios_capital_qtd": int(round(socios_capital_qtd)),
    }


@cache_data(show_spinner=False)
def get_extracted_data(ano: int, selected_months: list[int], file_mtime_ns: int | None = None) -> dict:
    """Extrai todos os dados necessários usando mapeamento exato da planilha Orçamento."""
    _ = file_mtime_ns
    f = str(ORCAMENTO_FILE)
    
    # --- 1. KPI AN (Margens) & 2. KPI AN (Pessoas) ---
    df_kpi = pd.read_excel(f, sheet_name="KPI AN", header=None, engine="openpyxl").fillna("")
    
    # Localizar a coluna "TOTAL" (última coluna)
    col_total_idx = None
    for j in range(len(df_kpi.columns)):
        if "TOTAL" in str(df_kpi.iloc[4, j]).upper():
            col_total_idx = j
            break
            
    mo, ml, pessoas = 0.0, 0.0, 0.0
    receita_kpi = 0.0
    resultado_kpi = 0.0
    
    if col_total_idx is not None:
        for i in range(len(df_kpi)):
            lbl = _normalize_text(df_kpi.iloc[i, 0])
            if "margem operacional" in lbl:
                mo = _to_float(df_kpi.iloc[i, col_total_idx])
            elif lbl == "receita":
                receita_kpi = _to_float(df_kpi.iloc[i, col_total_idx])
            elif lbl == "resultado":
                resultado_kpi = _to_float(df_kpi.iloc[i, col_total_idx])
            elif "quantidade de profissionais" in lbl:
                pessoas = _to_float(df_kpi.iloc[i, col_total_idx])
        
        if receita_kpi > 0:
            ml = resultado_kpi / receita_kpi

    # --- 3. Resumo -> Custos e Receita -> Mix de Receitas ---
    df_receita = pd.read_excel(f, sheet_name="Receita", header=None, engine="openpyxl").fillna("")
    receita_month_columns = _month_columns_from_header(df_receita, 7, ano, selected_months)
    contratuais = _sum_fixed_row(df_receita, 370, receita_month_columns)
    sucumb_1 = _sum_fixed_row(df_receita, 373, receita_month_columns)
    sucumb_2 = _sum_fixed_row(df_receita, 377, receita_month_columns)
    outras_linhas = [372, 374, 375, 376, 378, 379, 380]
    outras_receitas = sum(_sum_fixed_row(df_receita, row_index, receita_month_columns) for row_index in outras_linhas)
    sucumb = sucumb_1 + sucumb_2
    receita_total = contratuais + sucumb + outras_receitas
    totais_bi = _load_totais_period(selected_months, ano)
    if totais_bi["receita"]:
        receita_total = totais_bi["receita"]
        sucumb = totais_bi["sucumbencia"]
        outras_receitas = min(outras_receitas, max(receita_total - sucumb, 0.0))
        contratuais = max(receita_total - sucumb - outras_receitas, 0.0)
    pct_contratual = (contratuais / receita_total) if receita_total else 0.0
    pct_sucumbencia = (sucumb / receita_total) if receita_total else 0.0
    pct_outras = (outras_receitas / receita_total) if receita_total else 0.0

    performance_budget = load_performance_budget_metrics(str(PERFORMANCE_FILE), ano, selected_months)
    folha_data = _load_people_and_payroll_data(selected_months)
    omie_people_costs = load_people_costs_from_omie(str(OMIE_COSTS_FILE), selected_months)

    cost_totals = _extract_client_costs(f, selected_months)
    correspondentes = cost_totals["correspondentes"]
    outras_despesas = cost_totals["outras_despesas"]
    impostos = totais_bi["impostos"] if totais_bi["impostos"] else cost_totals["impostos"]
    if omie_people_costs["available"]:
        socios = omie_people_costs["partners"]
        socios_capital = 0.0
        clt = omie_people_costs["clt"]
        estagiarios = 0.0
        total_pessoas_custo = omie_people_costs["totalPeopleCost"]
    else:
        socios = folha_data["socios_val"]
        socios_capital = folha_data["socios_capital_val"]
        clt = folha_data["clt_val"]
        estagiarios = folha_data["estagiarios_val"]
        total_pessoas_custo = folha_data["total_pessoas_val"]
    total_despesas_detalhado = correspondentes + outras_despesas + impostos + total_pessoas_custo
    total_despesas_oficial = totais_bi["custos"]
    total_despesas = total_despesas_oficial
    cost_total_warning = ""
    if not total_despesas_oficial and total_despesas_detalhado:
        cost_total_warning = (
            "CUSTOS E DESPESAS veio zerado em INFORMAÇÕES GERENCIAIS.xlsx; "
            "o total oficial foi mantido como 0 e a OMIE ficou apenas no detalhe de pessoas."
        )

    receita_orcada_periodo = performance_budget["budgetRevenue"]
    meses_periodo = len(selected_months)
    meta_periodo_receita = receita_orcada_periodo
    pct_orcado = (receita_total / meta_periodo_receita) if meta_periodo_receita else 0.0

    receita_2025_periodo = totais_bi["receita_2025"]
    cost_total_anterior = totais_bi["custos_2025"]
    variacao_yoy = totais_bi["variacao_receita"]
    variacao_custos_yoy = totais_bi["variacao_custos"]
    if not receita_2025_periodo or not cost_total_anterior:
        try:
            df_comparativo = pd.read_excel(f, sheet_name="Comparativo 26-25", header=None, engine="openpyxl").fillna("")
            receita_2025_periodo = receita_2025_periodo or _get_comparativo_2025_period(df_comparativo, selected_months)
            cost_total_anterior = cost_total_anterior or _get_comparativo_total_2025_period(df_comparativo, "Despesa", selected_months)
        except Exception as exc:
            print(f"[ETL] Nao foi possivel ler Comparativo 26-25: {exc}")
        variacao_yoy = ((receita_total - receita_2025_periodo) / receita_2025_periodo) if receita_2025_periodo else 0.0
        variacao_custos_yoy = (
            (total_despesas - cost_total_anterior) / cost_total_anterior
            if cost_total_anterior
            else 0.0
        )

    # --- 4 & 5. Receita -> Faturamento por Cliente & Sucumbência ---
    clientes = []
    
    current_client = ""
    header_row_rec = -1
    
    for i in range(len(df_receita)):
        val0 = str(df_receita.iloc[i, 0]).strip()
        
        # Detecta início de bloco de cliente
        if "- FATURAMENTO" in val0.upper() and val0.upper() != "FATURAMENTO":
            current_client = val0.upper().replace("- FATURAMENTO", "").strip()
            # Procurar linha com datas logo abaixo
            for h in range(i+1, min(i+10, len(df_receita))):
                if df_receita.iloc[h, 1] != "" and df_receita.iloc[h, 1] != 0:
                    try:
                        if _is_date_in_period(df_receita.iloc[h, 1], ano, selected_months) or True:
                            header_row_rec = h
                            break
                    except:
                        pass
        
        # Extrai o TOTAL daquele bloco
        if val0.upper() == "TOTAL" and current_client and header_row_rec != -1:
            cliente_total = 0.0
            for j in range(1, len(df_receita.columns)):
                if _is_date_in_period(df_receita.iloc[header_row_rec, j], ano, selected_months):
                    cliente_total += _to_float(df_receita.iloc[i, j])
                    
            if cliente_total > 0 and current_client not in ["INTERCOMPANY"]:
                clientes.append((current_client, cliente_total))
            
            # Reseta as variáveis após fechar o bloco do cliente
            current_client = ""
            header_row_rec = -1

    clientes.sort(key=lambda x: x[1], reverse=True)
    top_5 = clientes[:5]
    outros = sum(x[1] for x in clientes[5:])
    
    ranking = [{"name": n, "value": _format_currency_full(v)} for n, v in top_5]
    if outros > 0:
        ranking.append({"name": "Outros", "value": _format_currency_full(outros)})

    total_card_costs = total_despesas
    item_sum = impostos + socios + socios_capital + clt + estagiarios + correspondentes + outras_despesas
    diff = total_card_costs - item_sum
    if abs(diff) > 0.01:
        print(f"[ETL] Diferença no card Estrutura de Custos: {diff:.2f}")
    
    return {
        "mo": mo,
        "ml": ml,
        "pessoas": folha_data["people"]["total"] or pessoas,
        "people": folha_data["people"],
        "receita_total": receita_total,
        "contratuais": contratuais,
        "sucumbencia": sucumb,
        "outras_receitas": outras_receitas,
        "pct_contratual": pct_contratual,
        "pct_sucumbencia": pct_sucumbencia,
        "pct_outras": pct_outras,
        "receita_orcada": receita_orcada_periodo,
        "receita_orcada_periodo": receita_orcada_periodo,
        "meta_periodo_receita": meta_periodo_receita,
        "meta_periodo_meses": meses_periodo,
        "pct_orcado": pct_orcado,
        "receita_2025_periodo": receita_2025_periodo,
        "variacao_yoy": variacao_yoy,
        "cost_total_anterior": cost_total_anterior,
        "variacao_custos_yoy": variacao_custos_yoy,
        "resultado_total": totais_bi["resultado"],
        "resultado_2025": totais_bi["resultado_2025"],
        "totais_bi": totais_bi,
        "budget": performance_budget,
        "people_costs": omie_people_costs,
        "sucumb": sucumb,
        "impostos": impostos,
        "top_5": ranking,
        "total_despesas": total_despesas,
        "total_despesas_oficial": total_despesas_oficial,
        "total_despesas_detalhado": total_despesas_detalhado,
        "cost_total_source": "INFORMAÇÕES GERENCIAIS.xlsx / TOTAIS  20 E 21 / CUSTOS E DESPESAS",
        "cost_total_warning": cost_total_warning,
        "socios_servico": socios,
        "socios_capital": socios_capital,
        "socios": socios,
        "clt": clt,
        "estagiarios": estagiarios,
        "total_pessoas_custo": total_pessoas_custo,
        "total_folha_val": folha_data["total_folha_val"],
        "folha_months_used": folha_data["folha_months_used"],
        "folha_snapshot_month": folha_data["folha_snapshot_month"],
        "correspondentes": correspondentes,
        "outras_despesas": outras_despesas,
        "clientes": clientes
    }


def get_dashboard_data(mes: int, ano: int, selected_months: list[int] | None = None) -> dict[str, Any]:
    """Orquestra o ETL a partir do Orçamento.xlsx e devolve a estrutura aninhada esperada pela interface."""
    mes = max(1, min(int(mes), 12))
    ano = int(ano)
    if not selected_months:
        selected_months = list(range(1, mes + 1))
        
    template = deepcopy(_load_dashboard_template(TEMPLATE_FILE.stat().st_mtime_ns))
    
    data = get_extracted_data(ano, selected_months, ORCAMENTO_FILE.stat().st_mtime_ns)
    
    receita_total = data["receita_total"]
    receita_orcada = data["receita_orcada"]
    meta_periodo_receita = data["meta_periodo_receita"]
    meta_periodo_meses = data["meta_periodo_meses"]
    pct_orcado = data["pct_orcado"]
    receita_2025_periodo = data["receita_2025_periodo"]
    variacao_yoy = data["variacao_yoy"]
    contratuais = data["contratuais"]
    sucumb = data["sucumbencia"]
    outras_receitas = data["outras_receitas"]
    pct_contratual = data["pct_contratual"]
    pct_sucumbencia = data["pct_sucumbencia"]
    pct_outras = data["pct_outras"]
    budget = data["budget"]
    people_costs = data["people_costs"]
    
    template["header"]["subtitle"] = f"Período acumulado: {_format_months_label(selected_months, ano)}"
    template["budget"] = {
        "revenue": budget["budgetRevenue"],
        "costs": budget["budgetCosts"],
        "result": budget["budgetResult"],
        "source": budget["source"],
        "logic": budget["budgetLogic"],
        "months": budget.get("months", []),
        "validation": budget.get("validation", {}),
        "error": budget.get("error"),
    }
    template["peopleCosts"] = {
        "partners": people_costs["partners"],
        "clt": people_costs["clt"],
        "total": people_costs["totalPeopleCost"],
        "breakdown": people_costs["breakdown"],
        "source": people_costs["source"],
        "months": people_costs.get("months", []),
        "available": people_costs.get("available", False),
        "error": people_costs.get("error"),
    }

    data_quality_warnings = []
    real_revenue_is_zero = abs(receita_total) < 0.01
    real_result_is_zero = abs(data["resultado_total"]) < 0.01
    zero_official_months = [
        item["month"]
        for item in data["totais_bi"].get("official_months", [])
        if abs(item.get("receita", 0.0)) < 0.01
        and abs(item.get("custos", 0.0)) < 0.01
        and abs(item.get("resultado", 0.0)) < 0.01
    ]
    revenue_breakdown_is_zero = (
        abs(contratuais) < 0.01
        and abs(sucumb) < 0.01
        and abs(outras_receitas) < 0.01
        and not data["top_5"]
    )
    official_cost_is_zero = abs(data["total_despesas_oficial"]) < 0.01

    if real_revenue_is_zero and budget["budgetRevenue"]:
        data_quality_warnings.append(
            "Receita real veio zerada na fonte oficial, mas existe orçamento no período."
        )
    if real_result_is_zero and budget["budgetResult"]:
        data_quality_warnings.append(
            "Resultado real veio zerado na fonte oficial, mas existe orçamento no período."
        )
    if revenue_breakdown_is_zero and budget["budgetRevenue"]:
        data_quality_warnings.append(
            "Abertura de receita e ranking de clientes vieram zerados/vazios para o período."
        )
    if data["cost_total_warning"]:
        data_quality_warnings.append(data["cost_total_warning"])
    if zero_official_months and (budget["budgetRevenue"] or budget["budgetCosts"] or budget["budgetResult"]):
        data_quality_warnings.append(
            f"Meses oficiais com receita/custos/resultado zerados: {zero_official_months}."
        )

    has_suspect_real_data = (
        (budget["budgetRevenue"] or budget["budgetCosts"] or budget["budgetResult"])
        and (
            (
                real_revenue_is_zero
                and real_result_is_zero
                and revenue_breakdown_is_zero
                and official_cost_is_zero
            )
            or bool(zero_official_months)
        )
    )

    template["dataQuality"] = {
        "hasSuspectRealData": bool(has_suspect_real_data),
        "warnings": data_quality_warnings,
        "zeroOfficialMonths": zero_official_months,
        "officialCostSource": data["cost_total_source"],
        "detailedCostTotal": data["total_despesas_detalhado"],
        "peopleCostSource": people_costs["source"] if people_costs.get("available") else "Resumo - folha.xlsx",
    }
    
    template["revenue_mix"] = {
        "icon": "💰",
        "title": "Estrutura de Receita",
        "value": receita_total,
        "contratual": contratuais,
        "sucumbencia": sucumb,
        "outras": outras_receitas,
        "outras_receitas": outras_receitas,
        "pct_contratual": pct_contratual,
        "pct_sucumbencia": pct_sucumbencia,
        "pct_outras": pct_outras,
        "pct_orcado": pct_orcado,
        "receita_orcada": receita_orcada,
        "receita_orcada_periodo": budget["budgetRevenue"],
        "meta_periodo_receita": meta_periodo_receita,
        "meta_periodo_meses": meta_periodo_meses,
        "receita_2025_periodo": receita_2025_periodo,
        "variacao_yoy": variacao_yoy,
        "subtitle": "",
        "comparison_pct": variacao_yoy * 100,
        "rows": [
            {
                "label": "Contratuais", 
                "value": _format_currency(contratuais), 
                "share": _format_percent(pct_contratual * 100)
            },
            {
                "label": "Sucumbência", 
                "value": _format_currency(sucumb), 
                "share": _format_percent(pct_sucumbencia * 100)
            },
            {
                "label": "Outras Receitas", 
                "value": _format_currency(outras_receitas), 
                "share": _format_percent(pct_outras * 100)
            }
        ],
        "expansions": [
            {
                "title": "Contratuais", 
                "items": [{"name": item["name"], "value": item["value"]} for item in data["top_5"][:5]]
            },
            {
                "title": "Sucumbência", 
                "items": []
            }
        ]
    }
    
    cost_total = data["total_despesas"]
    socios_total = data["socios_servico"] + data["socios_capital"]
    cost_groups = [
        {"label": "Impostos", "amount": data["impostos"]},
        {"label": "Sócios", "amount": data["socios"]},
        {"label": "CLT", "amount": data["clt"]},
        {"label": "Estagiários", "amount": data["estagiarios"]},
        {"label": "Correspondentes", "amount": data["correspondentes"]},
        {"label": "Outras Despesas", "amount": data["outras_despesas"]},
    ]
    cost_groups = [
        {"label": "Sócios", "amount": socios_total},
        {"label": "Impostos", "amount": data["impostos"]},
        {"label": "CLT", "amount": data["clt"]},
        {"label": "Estagiários", "amount": data["estagiarios"]},
        {"label": "Correspondentes", "amount": data["correspondentes"]},
        {"label": "Outras Despesas", "amount": data["outras_despesas"]},
    ]
    main_cost_group = max(cost_groups, key=lambda item: item["amount"])

    template["cost_structure"] = {
        "icon": "📉",
        "title": "Estrutura de Custos",
        "total": cost_total,
        "correspondentes": data["correspondentes"],
        "outras_despesas": data["outras_despesas"],
        "impostos": data["impostos"],
        "socios_servico": data["socios_servico"],
        "socios": socios_total,
        "clt": data["clt"],
        "estagiarios": data["estagiarios"],
        "total_pessoas_custo": data["total_pessoas_custo"],
        "official_total_costs": data["total_despesas_oficial"],
        "detailed_cost_total": data["total_despesas_detalhado"],
        "cost_total_source": data["cost_total_source"],
        "cost_total_warning": data["cost_total_warning"],
        "cost_total_anterior": data["cost_total_anterior"],
        "variacao_custos_yoy": data["variacao_custos_yoy"],
        "custo_orcado_periodo": budget["budgetCosts"],
        "meta_periodo_custos": budget["budgetCosts"],
        "pct_orcado_custos": (cost_total / budget["budgetCosts"]) if budget["budgetCosts"] else 0.0,
        "value": _format_currency(cost_total),
        "subtitle": "",
        "highlight": {
            "label": main_cost_group["label"],
            "value": _share(main_cost_group["amount"], cost_total),
            "caption": "principal grupo de custo"
        },
        "items": [
            {
                "label": "Impostos", 
                "value": _format_currency(data["impostos"]), 
                "share": _share(data["impostos"], cost_total), 
                "details": []
            },
            {
                "label": "Sócios", 
                "value": _format_currency(data["socios"]), 
                "share": _share(data["socios"], cost_total), 
                "details": []
            },
            {
                "label": "CLT", 
                "value": _format_currency(data["clt"]), 
                "share": _share(data["clt"], cost_total), 
                "details": []
            },
            {
                "label": "Estagiários", 
                "value": _format_currency(data["estagiarios"]), 
                "share": _share(data["estagiarios"], cost_total), 
                "details": []
            },
            {
                "label": "Correspondentes", 
                "value": _format_currency(data["correspondentes"]), 
                "share": _share(data["correspondentes"], cost_total), 
                "details": []
            },
            {
                "label": "Outras Despesas", 
                "value": _format_currency(data["outras_despesas"]), 
                "share": _share(data["outras_despesas"], cost_total), 
                "details": []
            }
        ]
    }

    template["cost_structure"].update({
        "socios_servico": data["socios_servico"],
        "socios_capital": data["socios_capital"],
        "socios": socios_total,
        "total_folha_val": data["total_folha_val"],
        "people_costs_source": people_costs["source"] if people_costs.get("available") else "Resumo - folha.xlsx",
        "folha_months_used": data["folha_months_used"],
        "folha_snapshot_month": data["folha_snapshot_month"],
        "items": [
            {
                "label": "Impostos",
                "value": _format_currency(data["impostos"]),
                "share": _share(data["impostos"], cost_total),
                "details": [],
            },
            {
                "label": "Sócios",
                "value": _format_currency(socios_total),
                "share": _share(socios_total, cost_total),
                "details": [],
            },
            {
                "label": "CLT",
                "value": _format_currency(data["clt"]),
                "share": _share(data["clt"], cost_total),
                "details": [],
            },
            {
                "label": "Estagiários",
                "value": _format_currency(data["estagiarios"]),
                "share": _share(data["estagiarios"], cost_total),
                "details": [],
            },
            {
                "label": "Correspondentes",
                "value": _format_currency(data["correspondentes"]),
                "share": _share(data["correspondentes"], cost_total),
                "details": [],
            },
            {
                "label": "Outras Despesas",
                "value": _format_currency(data["outras_despesas"]),
                "share": _share(data["outras_despesas"], cost_total),
                "details": [],
            },
        ],
    })
    
    template["net_result"] = {
        "icon": "📊",
        "title": "Resultado Líquido",
        "value": _format_currency(data["resultado_total"] if data["resultado_total"] else receita_total * data["ml"]),
        "subtitle": "",
        "resultado_orcado": budget["budgetResult"],
        "meta_periodo_resultado": budget["budgetResult"],
        "meta_periodo_meses": meta_periodo_meses,
        "pct_vs_orcado": ((data["resultado_total"] if data["resultado_total"] else receita_total * data["ml"]) / budget["budgetResult"]) if budget["budgetResult"] else 0.0,
        "resultado_2025": data["resultado_2025"],
        "variacao_2025": ((data["resultado_total"] - data["resultado_2025"]) / abs(data["resultado_2025"])) if data["resultado_2025"] else 0.0,
    }
    
    template["margins"] = {
        "icon": "📈",
        "title": "Margens",
        "value": _format_percent(data["mo"] * 100),
        "subtitle": "",
        "metrics": [
            {
                "label": "Margem Operacional", 
                "value": _format_percent(data["mo"] * 100), 
                "caption": ""
            },
            {
                "label": "Margem Líquida", 
                "value": _format_percent(data["ml"] * 100), 
                "caption": ""
            }
        ]
    }
    
    template["people"] = {
        "icon": "👥",
        "title": "Pessoas",
        "value": _format_count(data["people"]["total"]),
        "total": data["people"]["total"],
        "socios": data["people"]["socios"],
        "clt": data["people"]["clt"],
        "estagiarios": data["people"]["estagiarios"],
        "subtitle": "",
        "rows": [
            {
                "label": "Sócios",
                "value": _format_count(data["people"]["socios"]["qtd"]),
                "share": _format_percent(data["people"]["socios"]["pct"] * 100),
                "costValue": _format_currency(people_costs["partners"]) if people_costs.get("available") else "",
                "costSource": people_costs["source"] if people_costs.get("available") else "",
            },
            {
                "label": "CLT",
                "value": _format_count(data["people"]["clt"]["qtd"]),
                "share": _format_percent(data["people"]["clt"]["pct"] * 100),
                "costValue": _format_currency(people_costs["clt"]) if people_costs.get("available") else "",
                "costSource": people_costs["source"] if people_costs.get("available") else "",
            },
            {
                "label": "Estagiários",
                "value": _format_count(data["people"]["estagiarios"]["qtd"]),
                "share": _format_percent(data["people"]["estagiarios"]["pct"] * 100),
                "costValue": "",
                "costSource": "",
            },
        ]
    }

    template["people"]["rows"][0]["label"] = "Sócios de Serviço"
    
    template["top_clients"] = {
        "icon": "🏆",
        "title": "Top 5 Clientes",
        "subtitle": f"Faturamento em {_format_months_label(selected_months, ano)}",
        "ranking": data["top_5"]
    }
    
    template["insights"] = []
    template["technical_analysis"] = {
        "title": "Análise técnica",
        "paragraphs": []
    }
    
    metrics = build_dashboard_metrics(template)
    template["technical_analysis"] = generate_technical_analysis(template, metrics=metrics)
    template["insights"] = generate_insights(template, metrics=metrics)
    return template


def get_dashboard_content(
    selected_year: int | None = None,
    selected_months: list[int] | None = None,
) -> dict[str, Any]:
    """Carrega o dashboard usando o período selecionado como padrão."""
    today = date.today()
    ano = int(selected_year or today.year)
    months = selected_months if selected_months is not None else list(range(1, today.month + 1))
    return get_dashboard_data(max(months) if months else today.month, ano, selected_months=months)


def load_antecipacao_lucros() -> dict[str, Any]:
    """Carrega a aba de ajuste/antecipação mensal de lucros."""
    if not ANTECIPACAO_FILE.exists():
        return {
            "title": "AJUSTE / ANTECIPAÇÃO MENSAL DE DISTRIBUIÇÃO DE LUCROS",
            "subtitle": "Planilha de antecipação não encontrada.",
            "source": ANTECIPACAO_FILE.name,
            "summary": {},
            "gan": [],
            "gaa": [],
            "metrics": [],
            "companies": [],
            "totals": {},
            "missing_fields": [f"arquivo: {ANTECIPACAO_FILE.name}"],
        }

    companies_config = [
        {
            "sheet_name": "Antecipação Lucros GAN",
            "short_name": "GAN",
            "company_name": "GONDIM ADVOGADOS",
        },
        {
            "sheet_name": "Antecipação Lucros GAA",
            "fallback_sheet_name": "Antecipação Lucros GAT",
            "short_name": "GAA",
            "company_name": "GONDIM GESTÃO E TECNOLOGIA",
        },
    ]

    workbook = pd.ExcelFile(ANTECIPACAO_FILE)
    missing_fields: list[str] = []
    summary_sheet = "Antecipação Consolidado"
    summary = {
        "resultadoMensal": 0.0,
        "percentualAjusteMensalLucros": 0.0,
        "totalAjusteMensalLucros": 0.0,
        "totalQuotas": 0.0,
        "ajusteMensalPorQuota": 0.0,
    }

    if summary_sheet in workbook.sheet_names:
        df_summary = pd.read_excel(ANTECIPACAO_FILE, sheet_name=summary_sheet, header=None, engine="openpyxl").fillna("")
        summary_fields = [
            ("resultadoMensal", "Resultado Mensal", 0),
            ("percentualAjusteMensalLucros", "% Ajuste Mensal de Lucros", 1),
            ("totalAjusteMensalLucros", "Total de Ajuste Mensal de Lucros", 2),
            ("totalQuotas", "Total de Quotas", 3),
            ("ajusteMensalPorQuota", "Ajuste Mensal por Quota", 4),
        ]

        for key, label, row_index in summary_fields:
            if df_summary.shape[0] <= row_index or df_summary.shape[1] < 2:
                missing_fields.append(f"{summary_sheet}!B{row_index + 1} ({label})")
                continue
            summary[key] = _to_float(df_summary.iloc[row_index, 1])
    else:
        missing_fields.append(f"aba: {summary_sheet}")

    companies = []
    company_rows_by_short_name: dict[str, list[dict[str, Any]]] = {}
    total_adjustment = 0.0
    total_quotas = 0.0
    total_final = 0.0

    for config in companies_config:
        sheet_name = config["sheet_name"]
        if sheet_name not in workbook.sheet_names:
            sheet_name = config.get("fallback_sheet_name", sheet_name)
        if sheet_name not in workbook.sheet_names:
            missing_fields.append(f"aba: {config['sheet_name']}")
            continue

        df = pd.read_excel(ANTECIPACAO_FILE, sheet_name=sheet_name, header=0, engine="openpyxl").fillna("")
        rows = []
        company_base = 0.0
        company_adjustment = 0.0
        company_final = 0.0
        company_quotas = 0.0

        for index, row in df.iterrows():
            level = str(row.iloc[1]).strip()
            if not level:
                continue

            base = _to_float(row.iloc[2])
            quotas = _to_float(row.iloc[3])
            adjustment = _to_float(row.iloc[4])
            final = _to_float(row.iloc[5])

            company_base += base
            company_quotas += quotas
            company_adjustment += adjustment
            company_final += final

            rows.append({
                "index": int(index) + 1,
                "level": level,
                "nivelSocietario": level,
                "base": base,
                "quotas": quotas,
                "quotasServico": quotas,
                "adjustment": adjustment,
                "ajusteMensal": adjustment,
                "final": final,
                "antecipacaoFinal": final,
                "base_formatted": _format_currency(base),
                "adjustment_formatted": _format_currency(adjustment),
                "final_formatted": _format_currency(final),
                "quotas_formatted": _format_plain_decimal(quotas),
            })

        total_quotas += company_quotas
        total_adjustment += company_adjustment
        total_final += company_final

        companies.append({
            **config,
            "sheet_name": sheet_name,
            "rows": rows,
            "totals": {
                "base": company_base,
                "quotas": company_quotas,
                "adjustment": company_adjustment,
                "final": company_final,
                "base_formatted": _format_currency(company_base),
                "adjustment_formatted": _format_currency(company_adjustment),
                "final_formatted": _format_currency(company_final),
                "quotas_formatted": _format_plain_decimal(company_quotas),
            }
        })
        company_rows_by_short_name[config["short_name"].lower()] = rows

    summary_result = summary["resultadoMensal"]
    summary_adjustment_pct = summary["percentualAjusteMensalLucros"]
    summary_adjustment = summary["totalAjusteMensalLucros"]
    summary_quotas = summary["totalQuotas"]
    summary_adjustment_per_quota = summary["ajusteMensalPorQuota"]

    return {
        "title": "AJUSTE / ANTECIPAÇÃO MENSAL DE DISTRIBUIÇÃO DE LUCROS",
        "subtitle": f"Referência: {PROFIT_ADVANCE_REFERENCE} • Pagamento: {PROFIT_ADVANCE_PAYMENT}",
        "source": ANTECIPACAO_FILE.name,
        "periodoReferencia": PROFIT_ADVANCE_REFERENCE,
        "periodoPagamento": PROFIT_ADVANCE_PAYMENT,
        "summary": summary,
        "gan": company_rows_by_short_name.get("gan", []),
        "gaa": company_rows_by_short_name.get("gaa", []),
        "metrics": [
            {"label": "Resultado Mensal", "value": _format_currency_precise(summary_result), "raw_value": summary_result},
            {"label": "% Ajuste Mensal", "value": _format_percent(summary_adjustment_pct * 100), "raw_value": summary_adjustment_pct},
            {"label": "Total de Ajuste", "value": _format_currency_precise(summary_adjustment), "raw_value": summary_adjustment},
            {"label": "Total de Quotas", "value": _format_plain_decimal(summary_quotas), "raw_value": summary_quotas},
            {"label": "Ajuste por Quota", "value": _format_plain_decimal(summary_adjustment_per_quota, 4), "raw_value": summary_adjustment_per_quota},
        ],
        "companies": companies,
        "totals": {
            "result": summary_result,
            "adjustment_pct": summary_adjustment_pct,
            "adjustment": summary_adjustment,
            "quotas": summary_quotas,
            "adjustment_per_quota": summary_adjustment_per_quota,
            "table_final": total_final,
            "table_adjustment": total_adjustment,
            "table_quotas": total_quotas,
        },
        "missing_fields": missing_fields,
    }
